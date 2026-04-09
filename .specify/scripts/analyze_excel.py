"""
Analyze an .xlsm/.xlsx workbook and emit a Markdown report summarising:
  - sheets (with dimensions, visibility, tab color)
  - columns (inferred data types from first N data rows)
  - formulas (unique patterns, per sheet)
  - defined names / named ranges
  - data validations (dropdowns → potential lookup relationships)
  - likely relationships (column names that end with _id / Id / FK hints)
  - sample rows (first 3)
  - VBA macro presence (xlsm)

Usage: python analyze_excel.py <path/to/workbook.xlsm> <path/to/output.md>
"""
from __future__ import annotations

import re
import sys
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


SAMPLE_ROWS = 3
TYPE_SAMPLE_DEPTH = 50  # rows used for data-type inference
HEADER_SCAN_DEPTH = 6   # rows to scan when auto-detecting the header row

_IDENTIFIER_RE = re.compile(r"^[a-z][a-z0-9 ]*(?:_[a-z0-9 ]+)*(?:\s*\(.*\))?$")


def detect_header_row(ws: Worksheet, max_col: int) -> int:
    """Score rows 1..HEADER_SCAN_DEPTH and return the one most likely to contain
    column headers. Heuristic: count non-empty cells that look like snake_case
    identifiers (e.g. 'member_id', 'invoice_no', 'type_name (auto)').
    A row wins if it has ≥3 such cells AND ≥40% of its non-empty cells look like
    identifiers. Fall back to row 1 if no row scores well.
    """
    best_row = 1
    best_score = -1.0
    for r in range(1, HEADER_SCAN_DEPTH + 1):
        non_empty = 0
        id_like = 0
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or v == "":
                continue
            non_empty += 1
            if isinstance(v, str):
                s = v.strip().lower()
                if _IDENTIFIER_RE.match(s) and ("_" in s or s.endswith("id") or s.endswith("no")):
                    id_like += 1
        if non_empty == 0:
            continue
        ratio = id_like / non_empty
        # require a minimum of identifier-ish cells to beat row 1
        score = id_like + ratio
        if id_like >= 3 and ratio >= 0.4 and score > best_score:
            best_score = score
            best_row = r
    return best_row


def infer_type(values: list[Any]) -> str:
    non_null = [v for v in values if v is not None and v != ""]
    if not non_null:
        return "empty"
    types = set()
    for v in non_null:
        if isinstance(v, bool):
            types.add("bool")
        elif isinstance(v, int):
            types.add("int")
        elif isinstance(v, float):
            types.add("float")
        elif hasattr(v, "isoformat"):
            types.add("datetime")
        elif isinstance(v, str):
            s = v.strip()
            if re.fullmatch(r"[-+]?\d+", s):
                types.add("int-str")
            elif re.fullmatch(r"[-+]?\d*\.\d+", s):
                types.add("float-str")
            elif re.fullmatch(r"[\w.+-]+@[\w-]+\.[\w.-]+", s):
                types.add("email")
            elif re.fullmatch(r"\+?\d[\d\s\-()]{5,}", s):
                types.add("phone")
            else:
                types.add("text")
        else:
            types.add(type(v).__name__)
    if len(types) == 1:
        return next(iter(types))
    return "mixed(" + "|".join(sorted(types)) + ")"


def guess_relationship(col_name: str) -> str | None:
    if not isinstance(col_name, str):
        return None
    n = col_name.strip().lower()
    if not n:
        return None
    # obvious FK naming
    if n.endswith("_id") or n.endswith("id") and n != "id":
        base = n[:-3] if n.endswith("_id") else n[:-2]
        return f"FK → likely `{base}` entity"
    if n in ("member", "members", "member_ref", "member id", "memberid"):
        return "FK → Member"
    if n in ("event_ref", "event id", "eventid"):
        return "FK → Event"
    return None


def extract_vba_info(path: Path) -> dict[str, Any]:
    info: dict[str, Any] = {"has_vba": False, "modules": []}
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            if "xl/vbaProject.bin" in names:
                info["has_vba"] = True
            info["internal_files"] = sorted(
                [n for n in names if n.startswith("xl/") and n.endswith((".xml", ".bin"))]
            )
    except Exception as e:
        info["error"] = str(e)
    return info


def analyze_sheet(ws: Worksheet, ws_formula) -> dict[str, Any]:
    """Analyze a single sheet. ws is values-only; ws_formula exposes formulas."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    result: dict[str, Any] = {
        "title": ws.title,
        "state": ws.sheet_state,
        "dimensions": f"{ws.dimensions} ({max_row} rows × {max_col} cols)",
        "header_row": 1,
        "headers": [],
        "columns": [],
        "sample_rows": [],
        "formulas": [],
        "data_validations": [],
        "tab_color": None,
    }

    try:
        if ws.sheet_properties.tabColor is not None:
            result["tab_color"] = str(ws.sheet_properties.tabColor.rgb or "")
    except Exception:
        pass

    if max_row == 0 or max_col == 0:
        return result

    # --- auto-detect header row (scan rows 1..HEADER_SCAN_DEPTH) ---
    header_row = detect_header_row(ws, max_col)
    result["header_row"] = header_row

    headers = []
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        headers.append(cell.value)
    result["headers"] = headers

    # --- per-column type inference (data starts below header row) ---
    data_start = header_row + 1
    for col_idx, header in enumerate(headers, start=1):
        values = []
        for r in range(data_start, min(max_row, data_start + TYPE_SAMPLE_DEPTH - 1) + 1):
            values.append(ws.cell(row=r, column=col_idx).value)
        col_type = infer_type(values)
        non_null_count = sum(1 for v in values if v not in (None, ""))
        rel = guess_relationship(header) if header else None
        result["columns"].append(
            {
                "name": header,
                "index": col_idx,
                "type": col_type,
                "sampled_non_null": non_null_count,
                "relationship_hint": rel,
            }
        )

    # --- sample rows (first SAMPLE_ROWS data rows below header) ---
    for r in range(data_start, min(max_row, data_start + SAMPLE_ROWS - 1) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        result["sample_rows"].append(row_vals)

    # --- formulas (scan from formula-aware workbook) ---
    if ws_formula is not None:
        formulas: dict[str, list[str]] = defaultdict(list)
        scan_limit = min(max_row, 500)  # cap scan
        for row in ws_formula.iter_rows(
            min_row=1, max_row=scan_limit, min_col=1, max_col=max_col
        ):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    pattern = re.sub(r"\d+", "#", cell.value)
                    formulas[pattern].append(cell.coordinate)
        for pat, coords in sorted(formulas.items(), key=lambda x: -len(x[1]))[:15]:
            result["formulas"].append(
                {"pattern": pat, "count": len(coords), "example_cells": coords[:3]}
            )

    # --- data validations ---
    try:
        for dv in ws.data_validations.dataValidation:
            result["data_validations"].append(
                {
                    "type": dv.type,
                    "operator": dv.operator,
                    "formula1": dv.formula1,
                    "formula2": dv.formula2,
                    "ranges": str(dv.sqref) if dv.sqref else "",
                    "allowBlank": dv.allowBlank,
                }
            )
    except Exception as e:
        result["data_validations"].append({"error": str(e)})

    return result


def fmt_cell(v: Any) -> str:
    if v is None:
        return ""
    s = str(v)
    s = s.replace("|", "\\|").replace("\n", " ")
    if len(s) > 60:
        s = s[:57] + "..."
    return s


def render_markdown(path: Path, wb_data: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append(f"# Database Analysis — `{path.name}`")
    lines.append("")
    lines.append(
        "> Auto-generated by `.specify/scripts/analyze_excel.py`. Re-run after any "
        "change to the source workbook."
    )
    lines.append("")

    # --- overview ---
    lines.append("## Overview")
    lines.append("")
    lines.append(f"- **Source file**: `{path}`")
    lines.append(f"- **File size**: {path.stat().st_size:,} bytes")
    lines.append(f"- **Total sheets**: {len(wb_data['sheets'])}")
    visible = [s for s in wb_data["sheets"] if s["state"] == "visible"]
    hidden = [s for s in wb_data["sheets"] if s["state"] != "visible"]
    lines.append(f"- **Visible**: {len(visible)} | **Hidden/VeryHidden**: {len(hidden)}")
    vba = wb_data.get("vba", {})
    lines.append(f"- **Contains VBA macros**: {'**YES**' if vba.get('has_vba') else 'No'}")
    if wb_data.get("defined_names"):
        lines.append(f"- **Defined names**: {len(wb_data['defined_names'])}")
    lines.append("")

    # --- sheet index ---
    lines.append("## Sheet Index")
    lines.append("")
    lines.append("| # | Sheet | State | Dimensions | Columns | Formulas | Relationships |")
    lines.append("|---|-------|-------|------------|---------|----------|---------------|")
    for i, s in enumerate(wb_data["sheets"], 1):
        cols = len([c for c in s["columns"] if c["name"]])
        fcount = sum(f["count"] for f in s["formulas"])
        relcount = sum(1 for c in s["columns"] if c["relationship_hint"])
        lines.append(
            f"| {i} | `{s['title']}` | {s['state']} | {s['dimensions']} | "
            f"{cols} | {fcount} | {relcount} |"
        )
    lines.append("")

    # --- defined names ---
    if wb_data.get("defined_names"):
        lines.append("## Defined Names (Named Ranges)")
        lines.append("")
        lines.append("| Name | Refers To |")
        lines.append("|------|-----------|")
        for name, ref in wb_data["defined_names"]:
            lines.append(f"| `{fmt_cell(name)}` | `{fmt_cell(ref)}` |")
        lines.append("")

    # --- per-sheet sections ---
    for s in wb_data["sheets"]:
        lines.append(f"## Sheet: `{s['title']}`")
        lines.append("")
        lines.append(f"- **State**: {s['state']}")
        lines.append(f"- **Dimensions**: {s['dimensions']}")
        lines.append(f"- **Header row (auto-detected)**: row {s.get('header_row', 1)}")
        if s.get("tab_color"):
            lines.append(f"- **Tab color**: `{s['tab_color']}`")
        lines.append("")

        # columns
        named_cols = [c for c in s["columns"] if c["name"]]
        if named_cols:
            lines.append("### Columns")
            lines.append("")
            lines.append(
                "| # | Name | Inferred Type | Non-null (first 50) | Relationship Hint |"
            )
            lines.append(
                "|---|------|---------------|---------------------|-------------------|"
            )
            for c in named_cols:
                lines.append(
                    f"| {c['index']} | `{fmt_cell(c['name'])}` | {c['type']} | "
                    f"{c['sampled_non_null']} | "
                    f"{c['relationship_hint'] or '—'} |"
                )
            lines.append("")

        # sample rows
        if s["sample_rows"]:
            lines.append("### Sample Rows")
            lines.append("")
            header_line = "| " + " | ".join(
                f"`{fmt_cell(h)}`" if h else "—" for h in s["headers"]
            ) + " |"
            lines.append(header_line)
            lines.append("|" + "|".join(["---"] * len(s["headers"])) + "|")
            for row in s["sample_rows"]:
                lines.append("| " + " | ".join(fmt_cell(v) for v in row) + " |")
            lines.append("")

        # formulas
        if s["formulas"]:
            lines.append("### Formula Patterns (top 15)")
            lines.append("")
            lines.append("| Pattern (digits → `#`) | Count | Example Cells |")
            lines.append("|------------------------|-------|---------------|")
            for f in s["formulas"]:
                lines.append(
                    f"| `{fmt_cell(f['pattern'])}` | {f['count']} | "
                    f"{', '.join(f['example_cells'])} |"
                )
            lines.append("")

        # data validations
        if s["data_validations"]:
            lines.append("### Data Validations (dropdowns / lookups)")
            lines.append("")
            lines.append("| Type | Operator | Formula1 | Ranges |")
            lines.append("|------|----------|----------|--------|")
            for dv in s["data_validations"]:
                if "error" in dv:
                    lines.append(f"| ERROR | — | {fmt_cell(dv['error'])} | — |")
                    continue
                lines.append(
                    f"| {dv.get('type') or '—'} | {dv.get('operator') or '—'} | "
                    f"`{fmt_cell(dv.get('formula1'))}` | {fmt_cell(dv.get('ranges'))} |"
                )
            lines.append("")

        lines.append("---")
        lines.append("")

    # --- relationships summary ---
    lines.append("## Relationship Summary (heuristic)")
    lines.append("")
    lines.append(
        "Column names ending in `_id` / `Id` and other naming hints are grouped here. "
        "These are candidate foreign keys to confirm against real domain knowledge."
    )
    lines.append("")
    rel_rows = []
    for s in wb_data["sheets"]:
        for c in s["columns"]:
            if c["relationship_hint"]:
                rel_rows.append((s["title"], c["name"], c["type"], c["relationship_hint"]))
    if rel_rows:
        lines.append("| Sheet | Column | Type | Hint |")
        lines.append("|-------|--------|------|------|")
        for r in rel_rows:
            lines.append(
                f"| `{fmt_cell(r[0])}` | `{fmt_cell(r[1])}` | {r[2]} | {r[3]} |"
            )
    else:
        lines.append("_No obvious FK-style columns detected by heuristic._")
    lines.append("")

    # --- VBA info ---
    if vba:
        lines.append("## VBA / Macro Footprint")
        lines.append("")
        lines.append(f"- **Has VBA**: {vba.get('has_vba')}")
        if vba.get("internal_files"):
            lines.append("- **Internal xl/ files**:")
            for n in vba["internal_files"][:30]:
                lines.append(f"  - `{n}`")
            if len(vba["internal_files"]) > 30:
                lines.append(f"  - _(+{len(vba['internal_files']) - 30} more)_")
        lines.append("")

    lines.append("## Notes & Caveats")
    lines.append("")
    lines.append(
        "- Data types are inferred from the first 50 data rows per column; columns "
        "with rare non-default types may be misclassified."
    )
    lines.append(
        "- Relationship detection is naming-based only. Confirm with domain experts "
        "before modeling as foreign keys."
    )
    lines.append(
        "- Sample rows may include sensitive (PII / financial) data — treat this "
        "analysis file as internal per Constitution Principle I."
    )
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: analyze_excel.py <workbook> <output.md>", file=sys.stderr)
        return 1
    src = Path(sys.argv[1]).resolve()
    dst = Path(sys.argv[2]).resolve()
    if not src.exists():
        print(f"File not found: {src}", file=sys.stderr)
        return 2

    # values-only workbook (fast, resolved values)
    wb_values = load_workbook(src, data_only=True, read_only=False, keep_vba=False)
    # formula-aware workbook
    wb_formulas = load_workbook(src, data_only=False, read_only=False, keep_vba=False)

    wb_data: dict[str, Any] = {"sheets": []}

    for title in wb_values.sheetnames:
        ws = wb_values[title]
        ws_f = wb_formulas[title] if title in wb_formulas.sheetnames else None
        wb_data["sheets"].append(analyze_sheet(ws, ws_f))

    # defined names
    defined = []
    try:
        for dn in wb_values.defined_names:
            try:
                ref = wb_values.defined_names[dn].value
            except Exception:
                ref = "?"
            defined.append((dn, ref))
    except Exception:
        pass
    wb_data["defined_names"] = defined

    # VBA
    wb_data["vba"] = extract_vba_info(src)

    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_text(render_markdown(src, wb_data), encoding="utf-8")
    print(f"Wrote {dst} ({dst.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
