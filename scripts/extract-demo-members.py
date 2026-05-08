"""
Extract 20 demo members + their primary contacts from the SweCham Excel
workbook into a JSON file for `seed-demo-members.ts` to consume.

Why a Python intermediate step:
  - The repo has no JS/TS Excel parser as a dependency.
  - openpyxl is already used by `.specify/scripts/analyze_excel.py`.
  - Running Python at seed time keeps PII out of any committed TS file —
    the JSON output lands in `scripts/_demo-data/` which is gitignored.

Output schema (scripts/_demo-data/demo-members.json):
  {
    "tenantSlug": "swecham",
    "planYear": 2026,
    "rows": [
      {
        "companyName": "...",                        // members.company_name
        "country": "TH" | "SE" | "...",              // ISO-3166 alpha-2
        "taxId": "0105563084506" | null,             // members.tax_id (raw)
        "planId": "regular",                         // matches DB plan_id
        "registrationDate": "2025-04-04",            // ISO date
        "status": "active" | "inactive" | "archived",
        "notes": "Co-working space" | null,
        "primaryContact": {                          // optional
          "firstName": "Trinh",
          "lastName": "Danh",
          "email": "jin@theurbanoffice.com",
          "phone": null,                             // E.164 if available
          "preferredLanguage": "en"
        }
      }
    ]
  }

Usage:
  python scripts/extract-demo-members.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path("docs/SweCham_Database_Template_v11.xlsx")
OUTPUT = Path("scripts/_demo-data/demo-members.json")
TARGET_COUNT = 20

# Phase 6 review-round 2 CR2 + TD-S1 — schemaVersion + workbook
# header assertions. The TS-side `seed-demo-members.ts` parses the
# JSON via zod against `DEMO_SCHEMA_VERSION = 1`. If columns drift
# in the Excel template (rename, drop, reorder), the Python extractor
# silently produced empty/garbage rows that the seed then wrote to
# the live swecham tenant — invisible until a tier-filter test failed
# weeks later. The header assertion below fails LOUD on first run
# instead.
SCHEMA_VERSION = 1

# Excel `members` sheet header columns (row 4) — the order pinned at
# v11. Bump SCHEMA_VERSION here AND on the TS side when this changes.
EXPECTED_MEMBERS_HEADERS = [
    "member_id", "company_name", "display_name", "type_name", "type_id",
    "annual_fee", "status", "product_services", "source_basis", "join_date",
    "expiry_date", "notes", "created_at", "updated_at", "billing_address",
    "tax_id", "branch", "billing_email",
]
EXPECTED_CONTACTS_HEADERS = [
    "contact_id", "member_id", "company", "full_name", "title",
    "email", "phone", "is_primary", "role",
]


def assert_schema(ws, expected_headers: list[str], sheet_name: str) -> None:
    """Phase 6 review-round 2 CR2 — assert sheet header row matches the
    pinned schema version. Fails loud on first column drift instead of
    silently producing wrong JSON."""
    header_row = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    actual = [str(c).strip() if c else "" for c in header_row[: len(expected_headers)]]
    if actual != expected_headers:
        print(
            f"SCHEMA DRIFT in '{sheet_name}' sheet (workbook {WORKBOOK.name}, "
            f"SCHEMA_VERSION={SCHEMA_VERSION}):\n"
            f"  expected: {expected_headers}\n"
            f"  actual:   {actual}\n"
            f"Fix: bump SCHEMA_VERSION here AND in seed-demo-members.ts "
            f"(DEMO_SCHEMA_VERSION), then update column-index references "
            f"in main() before re-running.",
            file=sys.stderr,
        )
        sys.exit(2)

# Excel `type_name` → DB `plan_id` (see scripts/seed-swecham-2026-plans.ts)
PLAN_NAME_TO_ID = {
    "Platinum": "platinum",  # partnership tier
    "Gold": "gold",  # partnership tier
    "Premium": "premium",  # corporate tier
    "Large": "large",
    "Regular": "regular",
    "Start-up": "start-up",
    "Individual": "individual",
}

# Excel `status` → DB enum (memberStatusEnum: active|inactive|archived)
STATUS_MAP = {
    "Active": "active",
    "Pending": "active",
    "Rolling": "active",
    "Unpaid": "active",
    "Cancelled": "archived",
    "Suspended": "inactive",
}

# Tier diversity quota — pick this many of each so the demo screen shows
# the full plan ladder. Sum = 20.
TIER_QUOTA = {
    "platinum": 1,
    "gold": 2,
    "premium": 2,
    "large": 3,
    "regular": 8,
    "start-up": 2,
    "individual": 2,
}


def validate_thai_tax_id_checksum(tax_id: str) -> bool:
    """Mirror of `src/modules/members/domain/policies/thai-tax-id-checksum.ts`.

    Revenue Department algorithm:
      - 13 digits total
      - Weight digits 1..12 by [13,12,11,10,9,8,7,6,5,4,3,2]
      - check = (11 - (weighted_sum mod 11)) mod 10
      - compare against digit 13
    """
    if not re.fullmatch(r"\d{13}", tax_id):
        return False
    weights = [13, 12, 11, 10, 9, 8, 7, 6, 5, 4, 3, 2]
    s = sum(int(tax_id[i]) * weights[i] for i in range(12))
    return ((11 - (s % 11)) % 10) == int(tax_id[12])


# Tier categories used to know when Thai tax-id is REQUIRED to pass
# checksum (FR-009a: Corporate + Partnership tiers when country='TH').
# Individual / Thai-Alumni tiers don't require it.
CORPORATE_OR_PARTNERSHIP_PLAN_IDS = {
    "platinum",  # partnership
    "gold",  # partnership
    "diamond",  # partnership
    "premium",  # corporate
    "large",  # corporate
    "regular",  # corporate
    "start-up",  # corporate
}


def sanitise_tax_id(
    raw: str | None, country: str, plan_id: str
) -> tuple[str | None, str | None]:
    """Returns (sanitised_value, drop_reason).

    Strips whitespace + common separators. For TH + Corporate/Partnership
    tiers, runs the official 13-digit checksum and returns (None, reason)
    if it fails. For non-TH or Individual tiers, accepts up to 50 chars
    (matching the Domain `asTaxId()` rules)."""
    if not raw:
        return None, None
    cleaned = re.sub(r"[\s\-]", "", str(raw).strip())
    if not cleaned:
        return None, None
    if len(cleaned) > 50:
        return None, "too_long"

    if country == "TH" and plan_id in CORPORATE_OR_PARTNERSHIP_PLAN_IDS:
        if not re.fullmatch(r"\d{13}", cleaned):
            return None, "not_13_digits"
        if not validate_thai_tax_id_checksum(cleaned):
            return None, "bad_checksum"

    return cleaned, None


def detect_country(billing_address: str | None) -> str:
    """Heuristic: address contains a country name → ISO alpha-2.

    Default to 'TH' (most SweCham members are Thai-registered)."""
    if not billing_address:
        return "TH"
    a = billing_address.lower()
    # Order matters — most specific tokens first. Strong Thai signals win
    # before any Sweden fallback so we don't mis-classify Bangkok addresses
    # whose 5-digit zip happens to look like a Swedish postal code.
    if (
        "thailand" in a
        or "bangkok" in a
        or "phuket" in a
        or "chiang mai" in a
        or "samut " in a
        or "nonthaburi" in a
        or "pathum thani" in a
        or re.search(r"\bbangkok\s*\d{5}\b", a)
    ):
        return "TH"
    if "hong kong" in a or "kowloon" in a or " hk " in f" {a} ":
        return "HK"
    if "singapore" in a:
        return "SG"
    if "denmark" in a or "copenhagen" in a:
        return "DK"
    if "norway" in a or "oslo" in a:
        return "NO"
    if "finland" in a or "helsinki" in a:
        return "FI"
    if "vietnam" in a or "ho chi minh" in a or "hanoi" in a:
        return "VN"
    if "germany" in a:
        return "DE"
    if "united kingdom" in a or "london" in a:
        return "GB"
    if "united states" in a or " usa" in a:
        return "US"
    # Sweden last — addresses like "Sandvägen 4, 352 45 Växjö" or explicit
    # "Sweden". The 3+2 postal pattern alone is unreliable so we additionally
    # require either the country word or Swedish-specific characters.
    if "sweden" in a or re.search(r"[åäö]", a):
        return "SE"
    if re.search(r"\b\d{3}\s\d{2}\b\s[A-Za-zÅÄÖåäö]", billing_address):
        return "SE"
    return "TH"


def normalise_phone(raw: str | None) -> str | None:
    """Try to coerce raw phone strings into E.164. Drop on failure."""
    if not raw:
        return None
    s = re.sub(r"[\s\-()]", "", str(raw).strip())
    if not s:
        return None
    # If already E.164 (+digits), keep.
    if re.fullmatch(r"\+\d{8,15}", s):
        return s
    # Drop leading 0 and prepend +66 (Thailand is the dominant case).
    if re.fullmatch(r"0\d{8,9}", s):
        return "+66" + s[1:]
    # Already 10–15 digits without '+': prepend '+' if it looks reasonable.
    if re.fullmatch(r"\d{10,15}", s):
        return "+" + s
    return None


def split_full_name(full_name: str) -> tuple[str, str]:
    """Split 'Trinh Danh' → ('Trinh', 'Danh'). Single-word names go to firstName."""
    parts = full_name.strip().split()
    if len(parts) == 1:
        return parts[0], parts[0]  # repeat to satisfy `lastName NOT NULL`
    return parts[0], " ".join(parts[1:])


def looks_like_email(s: str | None) -> bool:
    return bool(s and re.fullmatch(r"[\w.+-]+@[\w-]+\.[\w.-]+", str(s).strip()))


def main() -> int:
    if not WORKBOOK.exists():
        print(f"workbook not found: {WORKBOOK}", file=sys.stderr)
        return 2

    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)

    # --- members sheet (header row = 4, data row 5..) ---
    ws_m = wb["members"]
    assert_schema(ws_m, EXPECTED_MEMBERS_HEADERS, "members")
    members: list[dict] = []
    for row in ws_m.iter_rows(min_row=5, max_row=ws_m.max_row or 5, values_only=True):
        if not row or len(row) < 18:
            continue
        # Phase 6 review-round 2 Simp5 — named indexed access aligned
        # with EXPECTED_MEMBERS_HEADERS above. Drift triggers
        # assert_schema() above before we ever reach this loop.
        member_id, company_name = row[0], row[1]
        type_name, status = row[3], row[6]
        product_services, join_date = row[7], row[9]
        billing_address, tax_id, billing_email = row[14], row[15], row[17]
        if not company_name or not member_id:
            continue
        plan_id = PLAN_NAME_TO_ID.get(str(type_name).strip()) if type_name else None
        if not plan_id:
            continue
        mapped_status = STATUS_MAP.get(str(status).strip(), "active") if status else "active"
        # Skip archived/inactive for a clean demo.
        if mapped_status != "active":
            continue
        reg_date = None
        if join_date and hasattr(join_date, "isoformat"):
            reg_date = join_date.date().isoformat()
        elif join_date:
            reg_date = str(join_date)[:10]
        country = detect_country(str(billing_address) if billing_address else None)
        sanitised_tax, drop_reason = sanitise_tax_id(
            str(tax_id) if tax_id else None, country, plan_id
        )
        members.append(
            {
                "_member_id": str(member_id).strip(),  # internal key for contact join
                "_tax_drop_reason": drop_reason,  # for summary stats only
                "companyName": str(company_name).strip(),
                "country": country,
                "taxId": sanitised_tax,
                "planId": plan_id,
                "registrationDate": reg_date or "2025-01-01",
                "status": mapped_status,
                "notes": (
                    str(product_services).strip()
                    if product_services and len(str(product_services).strip()) <= 500
                    else None
                ),
                "billingEmail": (
                    str(billing_email).strip() if looks_like_email(billing_email) else None
                ),
            }
        )

    # --- contacts sheet (header row = 4) — keyed by _member_id ---
    ws_c = wb["contacts"]
    assert_schema(ws_c, EXPECTED_CONTACTS_HEADERS, "contacts")
    primary_by_member: dict[str, dict] = {}
    for row in ws_c.iter_rows(min_row=5, max_row=ws_c.max_row or 5, values_only=True):
        # A contact_id, B member_id, C company, D full_name, E title,
        # F email, G phone, H is_primary, I role
        if not row or len(row) < 9:
            continue
        _, mref, _, full_name, title, email, phone, is_primary, _ = row[:9]
        if not mref or not full_name:
            continue
        mid = str(mref).strip()
        # Keep only the primary; if no primary, fall back to first contact seen.
        is_p = bool(is_primary) or str(is_primary).strip().upper() == "TRUE"
        if mid in primary_by_member and not is_p:
            continue
        if not looks_like_email(email):
            continue
        first, last = split_full_name(str(full_name))
        primary_by_member[mid] = {
            "firstName": first,
            "lastName": last,
            "email": str(email).strip().lower(),
            "phone": normalise_phone(str(phone) if phone else None),
            "roleTitle": str(title).strip() if title else None,
            "preferredLanguage": "en",
        }

    # --- pick 20 with tier diversity ---
    picked: list[dict] = []
    counts = {k: 0 for k in TIER_QUOTA}
    for m in members:
        plan = m["planId"]
        if counts.get(plan, 0) >= TIER_QUOTA.get(plan, 0):
            continue
        contact = primary_by_member.get(m["_member_id"])
        if not contact:
            continue  # require a primary contact for a meaningful demo
        picked.append(
            {
                "companyName": m["companyName"],
                "country": m["country"],
                "taxId": m["taxId"],
                "_tax_drop_reason": m.get("_tax_drop_reason"),
                "planId": m["planId"],
                "registrationDate": m["registrationDate"],
                "status": m["status"],
                "notes": m["notes"],
                "billingEmail": m["billingEmail"],
                "primaryContact": contact,
            }
        )
        counts[plan] += 1
        if len(picked) >= TARGET_COUNT:
            break

    # If quota under-filled (some tiers thin), top up with any remaining members.
    if len(picked) < TARGET_COUNT:
        seen_companies = {p["companyName"] for p in picked}
        for m in members:
            if m["companyName"] in seen_companies:
                continue
            contact = primary_by_member.get(m["_member_id"])
            if not contact:
                continue
            picked.append(
                {
                    "companyName": m["companyName"],
                    "country": m["country"],
                    "taxId": m["taxId"],
                    "_tax_drop_reason": m.get("_tax_drop_reason"),
                    "planId": m["planId"],
                    "registrationDate": m["registrationDate"],
                    "status": m["status"],
                    "notes": m["notes"],
                    "billingEmail": m["billingEmail"],
                    "primaryContact": contact,
                }
            )
            seen_companies.add(m["companyName"])
            if len(picked) >= TARGET_COUNT:
                break

    # Strip diagnostic-only field before serialising — keep summary aside
    # for the operator log so we can see how many tax IDs were dropped.
    drop_reasons: dict[str, int] = {}
    for p in picked:
        reason = p.pop("_tax_drop_reason", None)
        if reason:
            drop_reasons[reason] = drop_reasons.get(reason, 0) + 1

    payload = {
        "schemaVersion": SCHEMA_VERSION,
        "tenantSlug": "swecham",
        "planYear": 2026,
        "rows": picked,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    # Tier breakdown (no PII).
    breakdown: dict[str, int] = {}
    null_tax = 0
    for p in picked:
        breakdown[p["planId"]] = breakdown.get(p["planId"], 0) + 1
        if p["taxId"] is None:
            null_tax += 1
    print(f"wrote {OUTPUT} — {len(picked)} members")
    print("tier breakdown:", json.dumps(breakdown, sort_keys=True))
    print(f"tax_id null count: {null_tax}/{len(picked)}")
    if drop_reasons:
        print(f"tax_id drop reasons: {json.dumps(drop_reasons, sort_keys=True)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
